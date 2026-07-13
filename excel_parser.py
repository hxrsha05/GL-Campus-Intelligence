"""
Phase 2 — Excel Parser (header-aware)
Finds columns by scanning header names, not by fixed index.
Resilient to column insertions, reordering, and minor layout changes.
"""

import openpyxl
import re
import logging
from datetime import datetime
from pathlib import Path
from collections import defaultdict

log = logging.getLogger(__name__)

BASE_DIR = Path(__file__).parent


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_num(val, default=0):
    if val is None:
        return default
    if isinstance(val, str):
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def to_minutes(t) -> int:
    if t is None:
        return 0
    if hasattr(t, 'hour'):
        return t.hour * 60 + t.minute
    return 0


def normalize(s: str) -> str:
    """Lowercase, strip whitespace and newlines for fuzzy matching."""
    if not isinstance(s, str):
        return ""
    return " ".join(s.lower().split())


# Every name variant passed to find_sheet() so far. Populated as a side effect
# of parsing, so detect_unmatched_sheets() must be called AFTER the workbook
# has been fully parsed (see parse_electrical_file/parse_wtp_file) — it flags
# sheets that no parser call looked for, meaning their data was never read.
KNOWN_SHEET_VARIANTS: list = []

# Sheets confirmed (by manual inspection) to be vendor template scratch tabs
# or stale/legacy data with no bearing on current reporting. Silenced from
# detect_unmatched_sheets() so the alert email doesn't repeat every run for
# the same known-harmless sheets — it should only fire for something NEW.
#   - Sheet1..Sheet7, "Shift chart", "completed (2)": blank scratch tabs in
#     the vendor's workbook template, never contained real data.
#   - "RO water  reading" (two spaces): legacy sheet with readings dated
#     2012 — superseded by "RO water can count", which IS parsed.
IGNORED_SHEET_NAMES = {
    "sheet1", "sheet2", "sheet3", "sheet4", "sheet5", "sheet6", "sheet7",
    "shift chart", "completed (2)", "ro water reading",
    # Formerly read by parse_work_orders() for the old fake work-order KPI,
    # which was replaced by real Digii Tickets data from a separate source
    # (parse_ticket_file). Intentionally unused now, not a data gap.
    "electrical", "plumbing", "stp", "carpentry", "block-1&2 ac issue",
}


def find_sheet(wb, *name_variants: str):
    """Return the first sheet whose name loosely matches any variant."""
    KNOWN_SHEET_VARIANTS.extend(name_variants)
    norm_sheets = {normalize(n): n for n in wb.sheetnames}
    for variant in name_variants:
        key = normalize(variant)
        if key in norm_sheets:
            return wb[norm_sheets[key]]
    # fallback: partial match
    for variant in name_variants:
        key = normalize(variant)
        for norm_name, real_name in norm_sheets.items():
            if key in norm_name or norm_name in key:
                log.warning("Sheet fuzzy-matched '%s' -> '%s'", variant, real_name)
                return wb[real_name]
    raise KeyError(f"Sheet not found. Tried: {name_variants}. Available: {wb.sheetnames}")


# Sections that failed to parse this run (label -> error string), populated
# by _safe_section() and surfaced in each entry-point's result dict as
# "sectionErrors" so run_pipeline's freshness check can alert on it — a
# renamed/reworked sheet should blank ONE metric, not silently vanish with
# no signal at all (see _safe_section docstring for the failure this avoids).
SECTION_ERRORS: list = []


def _safe_section(label: str, fn):
    """
    Run a parser section (a lambda calling e.g. parse_eb_dg) and, if it fails
    with a KeyError (sheet not found — the case find_sheet() itself signals
    this way) or ValueError (e.g. no date rows found), log a warning and
    return None instead of propagating. Without this, one renamed/missing
    sheet raises out of parse_electrical_file() entirely, aborting injection
    for EVERY section in that file's data — not just the one sheet that
    changed. Records the failure in SECTION_ERRORS for the freshness check.
    """
    try:
        return fn()
    except (KeyError, ValueError) as e:
        msg = f"{label}: {e}"
        log.warning("Section failed to parse, skipping (dashboard section will show no new data): %s", msg)
        SECTION_ERRORS.append(msg)
        return None


def detect_unmatched_sheets(wb) -> list:
    """
    Return sheet names in this workbook that don't match (exactly or fuzzily)
    any name variant a parser has ever looked for. A non-empty result means
    either a brand-new sheet was added, or an existing one was renamed beyond
    what the fuzzy substring match in find_sheet() can recover — in both
    cases that sheet's data is silently NOT being parsed into the dashboard.
    """
    known_norm = {normalize(v) for v in KNOWN_SHEET_VARIANTS}
    unmatched = []
    for real_name in wb.sheetnames:
        norm_name = normalize(real_name)
        if norm_name in IGNORED_SHEET_NAMES:
            continue
        if norm_name in known_norm:
            continue
        if any(k in norm_name or norm_name in k for k in known_norm):
            continue
        unmatched.append(real_name)
    return unmatched


def find_columns(ws, keyword_map: dict, scan_rows: int = 6) -> dict:
    """
    Scan the first `scan_rows` rows to find column indices by keyword.
    keyword_map: { 'result_key': ['keyword1', 'keyword2', ...] }
    Returns: { 'result_key': col_index } — missing keys mean not found.
    """
    found = {}
    rows = list(ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True))

    for key, keywords in keyword_map.items():
        for row in rows:
            for col_idx, cell in enumerate(row):
                cell_norm = normalize(str(cell)) if cell is not None else ""
                for kw in keywords:
                    if normalize(kw) in cell_norm:
                        if key not in found:
                            found[key] = col_idx
                            break
                if key in found:
                    break

    missing = [k for k in keyword_map if k not in found]
    if missing:
        log.warning("Could not find columns for: %s", missing)

    return found


def find_data_start_row(ws, date_col: int, max_scan: int = 20) -> int:
    """Find the first row where the date column contains a datetime object."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), start=1):
        if date_col < len(row) and isinstance(row[date_col], datetime):
            return i
    return 5  # safe fallback


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def parse_eb_dg(wb, target_month: int, target_year: int) -> dict:
    ws = find_sheet(wb, "EB&DG Units", "EB&DG Units ")

    cols = find_columns(ws, {
        "eb":     ["ht eb units", "ht eb"],
        "occ":    ["students"],
        "dg1":    ["dg 1"],
        "dg2":    ["dg 2"],
        "dg3":    ["dg 3"],
        "dg_tot": ["total\n dg unit", "total dg unit"],
        "diesel": ["total\ndiesel consumption", "total diesel consumption"],
        "stock":  ["dg total tank stock", "total tank stock"],
    })

    date_col    = 0  # Date is always col A
    data_start  = find_data_start_row(ws, date_col)
    data        = defaultdict(lambda: {k: 0 for k in cols})

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[date_col] if row else None
        if not isinstance(cell_date, datetime):
            continue
        if cell_date.month != target_month or cell_date.year != target_year:
            continue
        day = cell_date.day
        # Unlike Solar/Buildings, "HT EB Units" here is a direct daily reading,
        # not a Final-Initial-Total-unit formula — so a not-yet-reported day
        # shows up as a genuinely blank cell rather than a negative formula
        # result. safe_num(None) would silently read as a valid 0 kWh day,
        # padding the count through the rest of the month, so check the raw
        # cell before treating this row as real data.
        eb_cidx = cols.get("eb")
        if eb_cidx is None or eb_cidx >= len(row) or not isinstance(row[eb_cidx], (int, float)):
            log.debug("EB&DG: skipping day %d — HT EB Units not yet entered", day)
            continue
        row_vals = {}
        for key, cidx in cols.items():
            if cidx < len(row):
                row_vals[key] = safe_num(row[cidx])
        # Skip formula-error rows — EB or DG negative means closing reading blank
        if row_vals.get("eb", 0) < 0 or row_vals.get("dg_tot", 0) < 0:
            log.debug("EB&DG: skipping day %d — negative value (incomplete row)", day)
            continue
        data[day].update(row_vals)

    if not data:
        log.warning("EB&DG: no rows found for %d/%d", target_month, target_year)
        return {}

    days = sorted(data.keys())
    return {
        "days":      days,
        "junEB":     [data[d].get("eb", 0)     for d in days],
        "junDG1":    [data[d].get("dg1", 0)    for d in days],
        "junDG2":    [data[d].get("dg2", 0)    for d in days],
        "junDG3":    [data[d].get("dg3", 0)    for d in days],
        "junDG":     [data[d].get("dg_tot", 0) for d in days],
        "junDiesel": [data[d].get("diesel", 0) for d in days],
        "junStock":  [data[d].get("stock", 0)  for d in days],
        "junOcc":    [data[d].get("occ", 0)    for d in days],
    }


def parse_power_cuts(wb, target_month: int, target_year: int) -> list:
    ws = find_sheet(wb, "EB power cut", "EB Power cut")

    cols = find_columns(ws, {
        "from":    ["from"],
        "to":      ["to"],
        "dur":     ["total time"],
        "remarks": ["remarks"],
    })

    date_col   = 0
    data_start = find_data_start_row(ws, date_col)
    cuts       = []
    cur_date   = None  # a day with multiple cuts only carries its date on the first row —
                        # subsequent cuts for that day are "continuation" rows with a blank date cell

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[date_col] if row else None
        if isinstance(cell_date, datetime):
            cur_date = cell_date
        if cur_date is None:
            continue
        if cur_date.month != target_month or cur_date.year != target_year:
            continue
        cell_date = cur_date

        dur_val  = row[cols["dur"]]    if "dur"     in cols and cols["dur"]     < len(row) else None
        from_val = row[cols["from"]]   if "from"    in cols and cols["from"]    < len(row) else None
        to_val   = row[cols["to"]]     if "to"      in cols and cols["to"]      < len(row) else None
        rem_val  = row[cols["remarks"]]if "remarks" in cols and cols["remarks"] < len(row) else None

        dur_mins = to_minutes(dur_val)
        if dur_mins == 0:
            continue

        start_str = f"{from_val.hour:02d}:{from_val.minute:02d}" if from_val and hasattr(from_val, 'hour') else "00:00"
        end_str   = f"{to_val.hour:02d}:{to_val.minute:02d}"     if to_val   and hasattr(to_val,   'hour') else "00:00"

        cuts.append({
            "date":   f"{cell_date.day:02d}-{cell_date.strftime('%b')}",
            "start":  start_str,
            "end":    end_str,
            "dur":    dur_mins,
            "events": f"{start_str}-{end_str}",
            "dg":     str(rem_val) if rem_val else "DG1",
        })

    log.info("Power cuts found: %d events", len(cuts))
    return cuts


def parse_power_cuts_all_months(wb) -> dict:
    """
    Same 'EB power cut' sheet as parse_power_cuts(), but instead of filtering
    to one target month, extracts EVERY month's individual outage events into
    {'YYYY-MM': [event, ...]}. The sheet is a continuous log (back to 2023),
    so history for past months (e.g. January's Longest Outage / Avg Cut,
    which need per-event detail, not just a monthly total) is sitting right
    there — it was just being discarded by the month filter. Keyed events use
    the same shape as parse_power_cuts()'s list so existing drill-down code
    (ddPowerCut) works unchanged against any month's list.
    """
    ws = find_sheet(wb, "EB power cut", "EB Power cut")

    cols = find_columns(ws, {
        "from":    ["from"],
        "to":      ["to"],
        "dur":     ["total time"],
        "remarks": ["remarks"],
    })

    date_col   = 0
    data_start = find_data_start_row(ws, date_col)
    by_month   = defaultdict(list)
    cur_date   = None

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[date_col] if row else None
        if isinstance(cell_date, datetime):
            cur_date = cell_date
        if cur_date is None:
            continue
        cell_date = cur_date

        dur_val  = row[cols["dur"]]    if "dur"     in cols and cols["dur"]     < len(row) else None
        from_val = row[cols["from"]]   if "from"    in cols and cols["from"]    < len(row) else None
        to_val   = row[cols["to"]]     if "to"      in cols and cols["to"]      < len(row) else None
        rem_val  = row[cols["remarks"]]if "remarks" in cols and cols["remarks"] < len(row) else None

        dur_mins = to_minutes(dur_val)
        if dur_mins == 0:
            continue

        start_str = f"{from_val.hour:02d}:{from_val.minute:02d}" if from_val and hasattr(from_val, 'hour') else "00:00"
        end_str   = f"{to_val.hour:02d}:{to_val.minute:02d}"     if to_val   and hasattr(to_val,   'hour') else "00:00"

        month_key = f"{cell_date.year:04d}-{cell_date.month:02d}"
        by_month[month_key].append({
            "date":   f"{cell_date.day:02d}-{cell_date.strftime('%b')}",
            "start":  start_str,
            "end":    end_str,
            "dur":    dur_mins,
            "events": f"{start_str}-{end_str}",
            "dg":     str(rem_val) if rem_val else "DG1",
        })

    log.info("Power cuts (all months) found: %d months, %d total events",
              len(by_month), sum(len(v) for v in by_month.values()))
    return dict(by_month)


def parse_rm_solar(wb, target_month: int, target_year: int) -> dict:
    ws = find_sheet(wb, "RM Solar panel reading", "RM Solar")

    cols = find_columns(ws, {
        "canteen":  ["canteen"],
        "hostel":   ["hostel lhs", "hostel"],
        "new_acad": ["new acad", "new acadamic"],
        "arr":      ["arr block", "arr"],
    }, scan_rows=5)

    # Sub-columns per section: [0]=Final reading, [2]=Total unit (offsets from section header)
    final_cols = {k: v for k, v in cols.items()}
    total_cols = {k: v + 2 for k, v in cols.items()}

    date_col   = 0
    data_start = find_data_start_row(ws, date_col)
    canteen, hostel, new_acad, arr, days = [], [], [], [], []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[date_col] if row else None
        if not isinstance(cell_date, datetime):
            continue
        if cell_date.month != target_month or cell_date.year != target_year:
            continue
        # A day with no "Final" reading entered yet isn't real data, even
        # though the Total-unit formula may evaluate it to 0 or a negative
        # number (blank treated as 0 in the subtraction) — checking Total
        # unit alone can't tell "not entered yet" apart from a real reading.
        finals = [row[final_cols[k]] if final_cols[k] < len(row) else None for k in final_cols]
        if not all(isinstance(v, (int, float)) for v in finals):
            log.debug("RM Solar: skipping day %d — Final reading not yet entered", cell_date.day)
            continue
        c = safe_num(row[total_cols["canteen"]]  if total_cols["canteen"]  < len(row) else None)
        h = safe_num(row[total_cols["hostel"]]   if total_cols["hostel"]   < len(row) else None)
        n = safe_num(row[total_cols["new_acad"]] if total_cols["new_acad"] < len(row) else None)
        a = safe_num(row[total_cols["arr"]]      if total_cols["arr"]      < len(row) else None)
        # Skip formula-error rows — closing reading blank causes negative values
        if c < 0 or h < 0 or n < 0 or a < 0:
            log.debug("RM Solar: skipping day %d — negative value (incomplete row)", cell_date.day)
            continue
        days.append(cell_date.day)
        canteen.append(c)
        hostel.append(h)
        new_acad.append(n)
        arr.append(a)

    log.info("RM Solar daily rows: %d", len(days))
    return {
        "rmSolarDays": days,
        "rmSolarDaily": {"canteen": canteen, "hostel": hostel, "newAcad": new_acad, "arr": arr},
    }


def parse_main_solar(wb, target_month: int, target_year: int) -> dict:
    ws = find_sheet(wb, "Solar Panel reading", "Solar panel reading")

    cols = find_columns(ws, {
        "pgdm":        ["pgdm hostel", "pgdm"],
        "new_acad":    ["new academic", "new acadamic"],
        "rhs":         ["pgpm hostel rhs", "rhs"],
        "bramahputra": ["bramahputra block", "bramahputra"],
        "admin_lhs":   ["admin lhs", "admin"],
    }, scan_rows=5)

    final_cols = {k: v for k, v in cols.items()}
    total_cols = {k: v + 2 for k, v in cols.items()}

    date_col   = 0
    data_start = find_data_start_row(ws, date_col)
    pgdm, new_acad, rhs, bramahputra, admin_lhs = [], [], [], [], []

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[date_col] if row else None
        if not isinstance(cell_date, datetime):
            continue
        if cell_date.month != target_month or cell_date.year != target_year:
            continue
        # A day with no "Final" reading entered yet isn't real data, even
        # though the Total-unit formula may evaluate it to 0 or a negative
        # number (blank treated as 0 in the subtraction) — checking Total
        # unit alone can't tell "not entered yet" apart from a real reading.
        finals = [row[final_cols[k]] if final_cols[k] < len(row) else None for k in final_cols]
        if not all(isinstance(v, (int, float)) for v in finals):
            log.debug("Main Solar: skipping day %d — Final reading not yet entered", cell_date.day)
            continue
        p  = safe_num(row[total_cols["pgdm"]]        if total_cols["pgdm"]        < len(row) else None)
        n  = safe_num(row[total_cols["new_acad"]]    if total_cols["new_acad"]    < len(row) else None)
        r  = safe_num(row[total_cols["rhs"]]         if total_cols["rhs"]         < len(row) else None)
        b  = safe_num(row[total_cols["bramahputra"]] if total_cols["bramahputra"] < len(row) else None)
        a  = safe_num(row[total_cols["admin_lhs"]]   if total_cols["admin_lhs"]   < len(row) else None)
        # Skip formula-error rows — negative values mean closing reading is blank
        if p < 0 or n < 0 or r < 0 or b < 0 or a < 0:
            log.debug("Main Solar: skipping day %d — negative value (incomplete row)", cell_date.day)
            continue
        pgdm.append(p)
        new_acad.append(n)
        rhs.append(r)
        bramahputra.append(b)
        admin_lhs.append(a)

    log.info("Main Solar daily rows: %d", len(pgdm))
    return {"mainSolarDaily": {"pgdm": pgdm, "newAcad": new_acad, "rhs": rhs,
                                "bramahputra": bramahputra, "adminLhs": admin_lhs}}


def parse_ticket_file(file_path: str) -> dict:
    """
    Parse the Digii Tickets service-request export (SERVICE_REQUEST sheet).
    One row per ticket, all-time cumulative history — no month filter, the
    whole file is the current picture (unlike the daily-rolling reports).

    Escalation level is encoded in 'Current Work Centre' as a suffix
    (_Admin_I / _Admin_II / _Admin_III / _Admin1 / _Admin for IT), which is
    the FINAL level a ticket reached before being handled — not a per-stage
    log. Department/problem-type is the sheet's 'Service Name' column.
    """
    path = Path(file_path)
    log.info("Parsing ticket file: %s", path.name)
    # This export's declared sheet dimensions are wrong (understate real
    # row/column count), which silently truncates iter_rows() in read-only
    # mode — load normally instead so the full sheet is actually walked.
    wb = openpyxl.load_workbook(str(path), data_only=True)

    try:
        ws = find_sheet(wb, "SERVICE_REQUEST", "SERVICE REQUEST", "Service Request")
    except KeyError as e:
        raise ValueError(f"Ticket file has no recognizable SERVICE_REQUEST sheet: {e}")

    header = next(ws.iter_rows(min_row=1, max_row=1, max_col=32, values_only=True))
    header_norm = [normalize(str(c)) if c is not None else "" for c in header]
    def col(*keywords, default):
        for i, c in enumerate(header_norm):
            if any(kw in c for kw in keywords):
                return i
        return default
    id_col     = col("request id", default=0)
    svc_col    = col("service name", default=1)
    name_col   = col("name of requester", default=3)
    status_col = col("status of request", default=5)
    reqdt_col  = col("request date", default=6)
    resdt_col  = col("resolved date", default=7)
    tat_col    = col("turn around time", default=13)
    wc_col     = col("current work centre", default=11)
    assign_col = col("assigned to", default=12)
    max_col = len(header)

    # Generic shared help-desk accounts that route many unrelated tickets —
    # excluded from recurrence detection since they don't identify one person
    # with one recurring physical problem.
    GENERIC_REQUESTERS = {"msh common", "sodexo help desk", "it support desk"}

    LEVEL_SUFFIXES = [
        ("_admin_iii", "Level 3"),
        ("_admin_ii",  "Level 2"),
        ("_admin_i",   "Level 1"),
        ("_admin1",    "Level 1"),
        ("_admin",     "Level 1"),
    ]

    def parse_tat_minutes(s):
        if not s or not isinstance(s, str) or s.strip() == "--":
            return None
        h = re.search(r"(\d+)\s*Hours?", s)
        m = re.search(r"(\d+)\s*Minutes?", s)
        if not h and not m:
            return None
        return (int(h.group(1)) if h else 0) * 60 + (int(m.group(1)) if m else 0)

    def parse_dt(s):
        if isinstance(s, datetime):
            return s
        if isinstance(s, str) and s.strip():
            try:
                return datetime.strptime(s.strip(), "%Y-%m-%d %H:%M:%S")
            except ValueError:
                return None
        return None

    dept_totals = defaultdict(int)
    level_totals = defaultdict(int)
    dept_by_level = defaultdict(lambda: defaultdict(int))
    total_tickets = 0
    pending_count = 0
    tat_values = []
    tickets = []
    dropped_count = 0
    dropped_work_centres = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=max_col, values_only=True):
        if not row:
            continue
        svc = row[svc_col] if svc_col < len(row) else None
        wc  = row[wc_col] if wc_col < len(row) else None
        if not svc or not wc:
            continue

        dept = str(svc).strip().rstrip(".")
        wc_norm = normalize(str(wc))
        level = next((lbl for suf, lbl in LEVEL_SUFFIXES if wc_norm.endswith(suf)), None)
        if level is None:
            # Unrecognized escalation-level suffix (e.g. Digii adds a new
            # tier, or changes the IT queue's naming) — this ticket vanishes
            # from every metric below. Unlike a missing sheet, this failure
            # mode has no natural exception to catch, so it must be counted
            # explicitly and surfaced, or it silently undercounts forever.
            dropped_count += 1
            dropped_work_centres.add(str(wc).strip())
            continue

        status_raw = row[status_col] if status_col < len(row) else None
        status = "Pending" if (status_raw and str(status_raw).strip().upper() == "SUBMITTED") else "Resolved"
        if status == "Pending":
            pending_count += 1

        tat_min = parse_tat_minutes(row[tat_col] if tat_col < len(row) else None)
        if tat_min is not None:
            tat_values.append(tat_min)

        req_date = row[reqdt_col] if reqdt_col < len(row) else None
        if isinstance(req_date, datetime):
            req_date_str = req_date.strftime("%Y-%m-%d")
        elif isinstance(req_date, str) and req_date.strip():
            # This export stores dates as plain text ("2026-07-07 23:52:12"),
            # not native Excel datetimes like the other source files.
            req_date_str = req_date.strip()[:10]
        else:
            req_date_str = None

        dept_totals[dept] += 1
        level_totals[level] += 1
        dept_by_level[dept][level] += 1
        total_tickets += 1

        name = row[name_col] if name_col < len(row) else None
        req_dt = parse_dt(row[reqdt_col] if reqdt_col < len(row) else None)
        res_dt = parse_dt(row[resdt_col] if resdt_col < len(row) else None)

        tickets.append({
            "id": row[id_col] if id_col < len(row) else None,
            "dept": dept,
            "status": status,
            "date": req_date_str,
            "tat": tat_min,
            "level": level,
            "assignee": str(row[assign_col]).strip() if assign_col < len(row) and row[assign_col] else None,
            "_name": str(name).strip() if name else None,
            "_req_dt": req_dt,
            "_res_dt": res_dt,
        })

    avg_tat = round(sum(tat_values) / len(tat_values)) if tat_values else None

    # ── Month-wise breakdown ────────────────────────────────────────────────
    # The sheet is all-time cumulative with no month filter, but the "current
    # month" picture (raised this month, resolved this month, by department)
    # is its own useful KPI — bucket every ticket by its Request Date's
    # "YYYY-MM" so the dashboard can show a this-month overview alongside the
    # all-time totals above, and so a future month picker has real data ready.
    month_totals = defaultdict(int)
    month_dept = defaultdict(lambda: defaultdict(int))
    month_pending = defaultdict(int)
    month_resolved = defaultdict(int)
    month_tat_values = defaultdict(list)
    for t in tickets:
        if not t["date"] or len(t["date"]) < 7:
            continue
        month_key = t["date"][:7]  # "YYYY-MM"
        month_totals[month_key] += 1
        month_dept[month_key][t["dept"]] += 1
        if t["status"] == "Pending":
            month_pending[month_key] += 1
        else:
            month_resolved[month_key] += 1
        if t["tat"] is not None:
            month_tat_values[month_key].append(t["tat"])

    ticket_months = {}
    for month_key, total in month_totals.items():
        tat_vals = month_tat_values.get(month_key) or []
        ticket_months[month_key] = {
            "total": total,
            "pending": month_pending.get(month_key, 0),
            "resolved": month_resolved.get(month_key, 0),
            "avgTatMin": round(sum(tat_vals) / len(tat_vals)) if tat_vals else None,
            "byDept": dict(month_dept[month_key]),
        }

    now = datetime.now()
    cur_month_key = f"{now.year:04d}-{now.month:02d}"
    ticket_this_month = ticket_months.get(cur_month_key, {
        "total": 0, "pending": 0, "resolved": 0, "avgTatMin": None, "byDept": {},
    })

    # ── Recurring issue detection ──────────────────────────────────────────
    # Flag cases where the SAME person's ticket for the SAME problem type was
    # marked resolved, then they filed another ticket for that same problem
    # type again within 3 days — a strong signal the original fix didn't
    # hold. Generic shared help-desk accounts are excluded since they route
    # many unrelated tickets and don't identify one recurring physical issue.
    RECURRENCE_WINDOW_DAYS = 3
    by_person_dept = defaultdict(list)
    for t in tickets:
        if not t["_name"] or not t["_req_dt"]:
            continue
        key_name = normalize(t["_name"])
        if key_name in GENERIC_REQUESTERS:
            continue
        by_person_dept[(t["_name"], t["dept"])].append(t)

    recurring = []
    for (name, dept), tix in by_person_dept.items():
        tix.sort(key=lambda t: t["_req_dt"])
        for i in range(1, len(tix)):
            prev, cur = tix[i - 1], tix[i]
            if prev["status"] != "Resolved" or not prev["_res_dt"]:
                continue
            gap_days = (cur["_req_dt"] - prev["_res_dt"]).total_seconds() / 86400
            if 0 <= gap_days <= RECURRENCE_WINDOW_DAYS:
                recurring.append({
                    "name": name,
                    "dept": dept,
                    "resolvedDate": prev["_res_dt"].strftime("%Y-%m-%d"),
                    "reopenedDate": cur["_req_dt"].strftime("%Y-%m-%d"),
                    "gapDays": round(gap_days, 1),
                    "priorId": prev["id"],
                    "newId": cur["id"],
                })

    recurring_by_dept = defaultdict(int)
    for r in recurring:
        recurring_by_dept[r["dept"]] += 1

    for t in tickets:
        del t["_name"], t["_req_dt"], t["_res_dt"]

    dropped_summary = []
    if dropped_count:
        dropped_summary.append(
            f"{dropped_count} ticket row(s) dropped — unrecognized work centre "
            f"suffix (not one of {[s for _, s in LEVEL_SUFFIXES]}): "
            f"{sorted(dropped_work_centres)}"
        )
        log.warning(dropped_summary[-1])

    result = {
        "ticketTotal": total_tickets,
        "ticketPending": pending_count,
        "ticketAvgTatMin": avg_tat,
        "ticketsByDept": dict(dept_totals),
        "ticketsByLevel": dict(level_totals),
        "ticketsDeptByLevel": {d: dict(lv) for d, lv in dept_by_level.items()},
        "ticketRows": tickets,
        "ticketRecurring": recurring,
        "ticketRecurringByDept": dict(recurring_by_dept),
        "ticketMonths": ticket_months,
        "ticketThisMonth": ticket_this_month,
        "ticketThisMonthKey": cur_month_key,
        "ticketDroppedCount": dropped_count,
        "ticketSectionErrors": dropped_summary,
    }
    log.info("Tickets parsed: %d total (%d pending), avg TAT %s min, %d recurring issues, by dept: %s",
              total_tickets, pending_count, avg_tat, len(recurring), dict(dept_totals))
    log.info("  This month (%s): %d tickets (%d pending, %d resolved)",
              cur_month_key, ticket_this_month["total"], ticket_this_month["pending"], ticket_this_month["resolved"])
    return result


# ── New WTP sheet parsers ─────────────────────────────────────────────────────

def parse_consume_sheet(wb, sheet_names: list, target_month: int, target_year: int,
                         consume_col: int = 3, unit: str = "L") -> dict:
    """
    Generic parser for open/close/consume sheets (Laundromat, Canteen flow,
    New well, Sathiya motor, STP-1/2).
    consume_col: 0-indexed column containing pre-computed daily consume value.
    unit: 'L' → divide by 1000 to get KL. 'KL' → use as-is.
    Returns {day: value_in_KL}.
    """
    try:
        ws = find_sheet(wb, *sheet_names)
    except KeyError:
        log.warning("Sheet not found: %s", sheet_names)
        return {}

    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        if consume_col >= len(row):
            continue
        raw = row[consume_col]
        val = safe_num(raw)
        if val < 0:
            continue
        kl = round(val / 1000, 2) if unit == "L" else round(val, 2)
        result[dt.day] = kl
    return result


def parse_ac_outlet(wb, target_month: int, target_year: int) -> dict:
    """
    Parse AC Outlet water consumption sheet.
    Columns: DATE | BLOCK1 | BLOCK2 | BLOCK3 | BLOCK5 | BLOCK6 | ARR | NALANDA | FACULTY | BHARATHI | GYM | MANIMEGALAI | TOTAL
    Values are direct daily Litres (not cumulative).
    Returns {day: {blocks: [...], total: float}}.
    """
    try:
        ws = find_sheet(wb, "AC Outlet water consumption", "AC Outlet")
    except KeyError:
        log.warning("AC Outlet sheet not found")
        return {}

    block_labels = ["Block 1","Block 2","Block 3","Block 5","Block 6","ARR Block","Nalanda","Faculty","Bharathi","Gym","Manimegalai"]
    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        blocks = [safe_num(row[i+1] if i+1 < len(row) else None) for i in range(len(block_labels))]
        total  = safe_num(row[12] if len(row) > 12 else None) or sum(blocks)
        result[dt.day] = {"blocks": blocks, "total": total}
    log.info("AC Outlet parsed %d days", len(result))
    return result


def parse_stp_flow(wb, target_month: int, target_year: int) -> dict:
    """
    Parse STP-1 & 2 Flow meter reading sheet.
    STP-1 consume: col 3; STP-2 consume: col 9. Both in Litres.
    """
    try:
        ws = find_sheet(wb, "STP-1 & 2 Flow meter reading", "STP-1 & 2 Flow")
    except KeyError:
        log.warning("STP flow sheet not found")
        return {}

    data_start = find_data_start_row(ws, 0)
    stp1, stp2 = {}, {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        v1 = safe_num(row[3] if len(row) > 3 else None)
        v2 = safe_num(row[9] if len(row) > 9 else None)
        if v1 >= 0:
            stp1[dt.day] = round(v1 / 1000, 2)
        if v2 >= 0:
            stp2[dt.day] = round(v2 / 1000, 2)
    log.info("STP Flow parsed %d days (STP1), %d days (STP2)", len(stp1), len(stp2))
    return {"stp1": stp1, "stp2": stp2}


def parse_well_levels(wb, target_month: int, target_year: int) -> dict:
    """
    Parse Well motor water level sheet.
    Values are strings like '10 feet' — strip unit and cast to int.
    Returns {day: {acre2: int, acre6: int}}.
    """
    try:
        ws = find_sheet(wb, "Well motor water level", "Well motor")
    except KeyError:
        log.warning("Well motor water level sheet not found")
        return {}

    def parse_feet(val):
        if val is None:
            return None
        s = str(val).lower().replace("feet", "").replace("ft", "").strip()
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return None

    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        a2 = parse_feet(row[1] if len(row) > 1 else None)
        a6 = parse_feet(row[2] if len(row) > 2 else None)
        if a2 is not None or a6 is not None:
            result[dt.day] = {"acre2": a2 or 0, "acre6": a6 or 0}
    log.info("Well levels parsed %d days", len(result))
    return result


def parse_tds(wb, target_month: int, target_year: int) -> dict:
    """
    Parse TDS reading sheet.
    Layout: Date | S1_col1 | S1_col2 | S1_name | S2_col1 | S2_col2 | S2_name | S3_col1 | S3_col2 | S3_name
            | spacer | same pattern for RO-2000
    Data rows start at row 4.
    Returns {day: {ro1000_avg: float, ro2000_avg: float}}.
    """
    try:
        ws = find_sheet(wb, "TDS reading(Ro 1000) & (RO2000)", "TDS reading")
    except KeyError:
        log.warning("TDS sheet not found")
        return {}

    result = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        # RO-1000: cols 1,2,4,5,7,8 (3 shifts × 2 readings each)
        ro1000_vals = [safe_num(row[c] if len(row) > c else None)
                       for c in [1, 2, 4, 5, 7, 8]
                       if isinstance(row[c] if len(row) > c else None, (int, float))]
        # RO-2000: cols 10,11,13,14,16,17 (no spacer col before this block, unlike RO-1000)
        ro2000_vals = [safe_num(row[c] if len(row) > c else None)
                       for c in [10, 11, 13, 14, 16, 17]
                       if isinstance(row[c] if len(row) > c else None, (int, float))]
        ro1_avg = round(sum(ro1000_vals) / len(ro1000_vals), 1) if ro1000_vals else 0
        ro2_avg = round(sum(ro2000_vals) / len(ro2000_vals), 1) if ro2000_vals else 0
        result[dt.day] = {"ro1000": ro1_avg, "ro2000": ro2_avg}
    log.info("TDS parsed %d days", len(result))
    return result


def parse_ro_cans(wb, target_month: int, target_year: int) -> dict:
    """
    Parse RO water can count sheet.
    Col 1 = daily count (Nos).
    Returns {day: count}.
    """
    try:
        ws = find_sheet(wb, "RO water can count", "RO water can")
    except KeyError:
        log.warning("RO can count sheet not found")
        return {}

    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        val = safe_num(row[1] if len(row) > 1 else None)
        if val > 0:
            result[dt.day] = int(val)
    log.info("RO cans parsed %d days", len(result))
    return result


def parse_ev_meter(wb, target_month: int, target_year: int) -> dict:
    """
    Parse EV meter sheet.
    Headers at row 1, data at row 3+.
    Col 3 = Total unit (Final - Initial), already computed.
    """
    try:
        ws = find_sheet(wb, "EV meter")
    except KeyError:
        log.warning("EV meter sheet not found")
        return {}

    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        val = safe_num(row[3] if len(row) > 3 else None)
        if val > 0:
            result[dt.day] = round(val, 1)
    log.info("EV meter parsed %d days", len(result))
    return result


def parse_tenant_meter(wb, *sheet_name_variants: str, target_month: int, target_year: int) -> dict:
    """
    Parse a tenant sub-meter sheet (Nescafe, Chennai Beverage, etc).
    Header at row 1/2, data at row 3+. Col 3 = Total unit.
    Handles negative values (meter reset) by treating as 0.
    """
    try:
        ws = find_sheet(wb, *sheet_name_variants)
    except KeyError:
        log.warning("Tenant sheet not found: %s", sheet_name_variants)
        return {}

    data_start = find_data_start_row(ws, 0)
    result = {}
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        val = safe_num(row[3] if len(row) > 3 else None)
        if val > 0:
            result[dt.day] = round(val, 1)
    return result


def parse_laundry_elec(wb, target_month: int, target_year: int) -> dict:
    """
    Parse Laundromat reading (electrical consumption) sheet.
    Headers at row 4, data from row 5. Col 3 = Total unit.
    """
    try:
        ws = find_sheet(wb, "Laundromat reading", "Laundromat reading ")
    except KeyError:
        log.warning("Laundromat reading sheet not found")
        return {}

    # Data starts after 4 header rows
    result = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        val = safe_num(row[3] if len(row) > 3 else None)
        if val > 0:
            result[dt.day] = round(val, 1)
    log.info("Laundry elec parsed %d days", len(result))
    return result


# ── WTP / Water file parser ───────────────────────────────────────────────────

def parse_wtp_file(file_path: str, target_month: int = None, target_year: int = None) -> dict:
    """
    Parse the WTP water Excel file.
    Extracts daily WTP-1/2/3 consumption (KL), new well pump (KL),
    laundry (KL), and canteen RO (KL).
    Only includes rows where the closing reading is non-empty AND consume > 0
    to skip formula-error rows (e.g. the open row on the last day).
    """
    path = Path(file_path)
    log.info("Parsing WTP file: %s", path.name)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # ── Auto-detect month/year from the WTP sheet ─────────────────────────────
    try:
        wtp_ws = find_sheet(wb, "TPs", "wtp", "WTP's", "30 TPs", "30 TPs")
    except KeyError as e:
        raise ValueError(f"WTP file has no recognizable main sheet (tried 'TPs'/'wtp'/etc): {e}")
    if target_month is None or target_year is None:
        latest = None
        for row in wtp_ws.iter_rows(min_row=3, max_row=5000, values_only=True):
            if row and isinstance(row[0], datetime):
                latest = row[0]
        if latest:
            target_month = target_month or latest.month
            target_year  = target_year  or latest.year
            log.info("WTP auto-detected period: %s %d", latest.strftime("%B"), target_year)
        else:
            raise ValueError("Could not detect month from WTP Excel file")

    # ── WTP-1 / WTP-2 / WTP-3 from the main TPs sheet ────────────────────────
    # Sheet layout (from screenshot):
    #   Col A: Date
    #   Col B: WTP-1 Open (KiloLit)   Col C: WTP-1 Close   Col D: WTP-1 Consume(Lit)
    #   Col E: WTP-2 Open              Col F: WTP-2 Close   Col G: WTP-2 Consume(Lit)
    #   Col H: WTP-3 Open              Col I: WTP-3 Close   Col J: WTP-3 Consume(Lit)
    # Consume values are in Litres — divide by 1000 to get KL.

    wtp_by_day = {}  # {day: {wtp1, wtp2, wtp3}}

    data_start = find_data_start_row(wtp_ws, 0)
    for row in wtp_ws.iter_rows(min_row=data_start, values_only=True):
        if not row or not isinstance(row[0], datetime):
            continue
        dt = row[0]
        if dt.month != target_month or dt.year != target_year:
            continue
        day = dt.day

        # Closing reading must be present (not None/empty) to avoid formula-error rows
        close1 = row[2] if len(row) > 2 else None
        close2 = row[5] if len(row) > 5 else None
        close3 = row[8] if len(row) > 8 else None  # col I (0-indexed = 8)

        c1 = safe_num(row[3] if len(row) > 3 else None)   # WTP-1 Consume Lit
        c2 = safe_num(row[6] if len(row) > 6 else None)   # WTP-2 Consume Lit
        c3 = safe_num(row[9] if len(row) > 9 else None)   # WTP-3 Consume Lit  (col J = idx 9... but col K=10 in screenshot)

        # Skip rows where closing is blank or consume is negative (formula error)
        if close1 is None and close2 is None and close3 is None:
            continue
        if c1 < 0 or c2 < 0 or c3 < 0:
            log.debug("WTP: skipping day %d — negative consume (incomplete row)", day)
            continue

        wtp_by_day[day] = {
            "wtp1": round(c1 / 1000, 1) if c1 > 0 else 0,
            "wtp2": round(c2 / 1000, 1) if c2 > 0 else 0,
            "wtp3": round(c3 / 1000, 1) if c3 > 0 else 0,
        }

    wtp_days = sorted(wtp_by_day.keys())
    log.info("WTP daily rows parsed: %d days", len(wtp_days))

    # ── New well pump from "New well reading flow meter" sheet ────────────────
    well_by_day = {}
    try:
        well_ws = find_sheet(wb, "New well reading flow meter", "New well", "New Well")
        well_cols = find_columns(well_ws, {
            "consume": ["consume", "total", "kl"],
        })
        well_dcol = 0
        well_start = find_data_start_row(well_ws, well_dcol)
        ccol = well_cols.get("consume", 3)  # fallback col D

        for row in well_ws.iter_rows(min_row=well_start, values_only=True):
            if not row or not isinstance(row[0], datetime):
                continue
            dt = row[0]
            if dt.month != target_month or dt.year != target_year:
                continue
            val = safe_num(row[ccol] if ccol < len(row) else None)
            if val > 0:
                well_by_day[dt.day] = round(val / 1000, 1) if val > 1000 else val
    except KeyError:
        log.warning("New well sheet not found — skipping")

    # ── Laundromat water from dedicated Laundromat sheet ─────────────────────
    laundry_by_day = parse_consume_sheet(wb, ["Laundromat", "Laundromat "], target_month, target_year, consume_col=3, unit="L")

    # ── Canteen flow meter ────────────────────────────────────────────────────
    canteen_flow = parse_consume_sheet(wb, ["Canteen reading flow meter", "Canteen flow"], target_month, target_year, consume_col=3, unit="L")

    # ── Sathiya motor ─────────────────────────────────────────────────────────
    sathiya_by_day = parse_consume_sheet(wb, ["Sathiya motor", "Sathiya Motor"], target_month, target_year, consume_col=3, unit="L")

    # ── AC Outlet water consumption ───────────────────────────────────────────
    ac_outlet_by_day = parse_ac_outlet(wb, target_month, target_year)

    # ── STP-1 & STP-2 flow meters ─────────────────────────────────────────────
    stp_flow = parse_stp_flow(wb, target_month, target_year)

    # ── Well motor water levels ───────────────────────────────────────────────
    well_levels = parse_well_levels(wb, target_month, target_year)

    # ── TDS readings ──────────────────────────────────────────────────────────
    tds_by_day = parse_tds(wb, target_month, target_year)

    # ── RO water can count ────────────────────────────────────────────────────
    ro_cans = parse_ro_cans(wb, target_month, target_year)

    unmatched_sheets = detect_unmatched_sheets(wb)
    if unmatched_sheets:
        log.warning("Unrecognized sheet(s) in %s (not parsed): %s", path.name, unmatched_sheets)

    wb.close()

    # ── Build ordered day-indexed arrays ─────────────────────────────────────
    all_days = sorted(set(wtp_days) | set(well_by_day.keys()))
    if not all_days:
        all_days = wtp_days

    # Sathiya monthly total (for existing sathiyaMonthly array injection)
    sathiya_total = sum(sathiya_by_day.values())

    # AC outlet: per-block arrays and total (use all_days for alignment)
    ac_outlet_blocks = [
        [ac_outlet_by_day.get(d, {}).get("blocks", [0]*11)[i] for d in all_days]
        for i in range(11)
    ]
    ac_outlet_total = [ac_outlet_by_day.get(d, {}).get("total", 0) for d in all_days]

    return {
        "month":          target_month,
        "year":           target_year,
        "wtpDays":        wtp_days,
        "wtpDaily": {
            "wtp1": [wtp_by_day.get(d, {}).get("wtp1", 0) for d in wtp_days],
            "wtp2": [wtp_by_day.get(d, {}).get("wtp2", 0) for d in wtp_days],
            "wtp3": [wtp_by_day.get(d, {}).get("wtp3", 0) for d in wtp_days],
        },
        "newWellJunKL":   [well_by_day.get(d, 0)    for d in wtp_days],
        "laundryDaily":   [laundry_by_day.get(d, 0) for d in wtp_days],
        # New water data
        "canteenFlowDaily":  [canteen_flow.get(d)    for d in all_days],
        "sathiyaDailyKL":    [sathiya_by_day.get(d, 0)  for d in all_days],
        "sathiyaJunTotal":   sathiya_total,
        "acOutletTotal":     ac_outlet_total,
        "acOutletBlocks":    ac_outlet_blocks,
        "stpFlow1":          [stp_flow.get("stp1", {}).get(d, 0) for d in all_days],
        "stpFlow2":          [stp_flow.get("stp2", {}).get(d, 0) for d in all_days],
        "wellAcre2":         [well_levels.get(d, {}).get("acre2", None) for d in all_days],
        "wellAcre6":         [well_levels.get(d, {}).get("acre6", None) for d in all_days],
        "tdsRo1000":         [tds_by_day.get(d, {}).get("ro1000", None) for d in all_days],
        "tdsRo2000":         [tds_by_day.get(d, {}).get("ro2000", None) for d in all_days],
        "roCansDaily":       [ro_cans.get(d, 0) for d in all_days],
        "allDays":           all_days,
        "unmatchedSheets":   unmatched_sheets,
    }


# ── Building sub-meter panel sheets ────────────────────────────────────────────

BUILDING_SHEETS = [
    ("substation",    "24 hrs Reading - Substation", "Substation (24hr Panel)"),
    ("mainAcademic",  "Main Academic LHS & RHS",      "Main Academic LHS & RHS"),
    ("nalanda",       "Nalanda Block",                "Nalanda Block"),
    ("oldHostel",     "Old Hostel LHS & RHS",          "Old Hostel LHS & RHS"),
    ("pgdmHostel",    "PGDM Hostel",                   "PGDM Hostel"),
    ("canteen",       "Canteen",                       "Canteen"),
    ("canteenAddOn",  "Canteen Add on Panel",          "Canteen Add-On Panel"),
    ("arrBlock",      "ARR Block",                     "ARR Block"),
    ("thiruvalluvar", "Thiruvalluvar block",           "Thiruvalluvar Block"),
    ("saraswati",     "Saraswati block",               "Saraswati Block"),
    ("pumpsMotor",    "Pumps& Motor energy consumption", "Pumps & Motor Energy"),
    ("mainLighting",  "Main lighting panel",          "Main Lighting Panel"),
]


def _find_panel_columns(rows: list) -> tuple:
    """
    Given the first ~8 rows of a building panel sheet, locate the Final/Intial/
    Total-unit triad header row and return (panel_labels_by_total_col, triad_row_idx).
    Panel name rows sit either 1 or 2 rows above the triad row depending on
    whether a room-description row is present; pick whichever is shorter/more
    distinct, since description rows tend to be long free text.
    """
    triad_row_idx = None
    for i, row in enumerate(rows):
        norm = [normalize(str(c)) for c in row if c is not None]
        if "final" in norm and "intial" in norm:
            triad_row_idx = i
            break
    if triad_row_idx is None:
        return None, None

    triad_row = rows[triad_row_idx]
    total_cols = [ci for ci, c in enumerate(triad_row)
                  if c is not None and "total unit" in normalize(str(c))]
    if not total_cols:
        return None, None

    def labels_for(row_idx):
        if row_idx < 0:
            return []
        row = rows[row_idx]
        labels, last = [], None
        for ci, cell in enumerate(row):
            if cell is not None and str(cell).strip():
                last = str(cell).strip()
            if ci in total_cols:
                labels.append(last)
        return labels

    idx_near = triad_row_idx - 1
    labels = labels_for(idx_near)
    distinct_near = len({l for l in labels if l})
    avg_len = (sum(len(l) for l in labels if l) / max(1, sum(1 for l in labels if l)))

    idx_far = triad_row_idx - 2
    if idx_far >= 0 and (avg_len > 40 or distinct_near < len(total_cols)):
        far_labels = labels_for(idx_far)
        if len({l for l in far_labels if l}) > distinct_near:
            labels = far_labels

    return list(zip(labels, total_cols)), triad_row_idx


def _label_columns(panel_cols: list, fallback_label: str) -> list:
    """Dedup panel names sharing the same day-value (e.g. Gents hostel spanning 3 panels)."""
    seen = {}
    for name, col in panel_cols:
        label = name or fallback_label
        seen[label] = seen.get(label, 0) + 1
    counters = {}
    labeled_cols = []
    for name, col in panel_cols:
        label = name or fallback_label
        counters[label] = counters.get(label, 0) + 1
        if seen[label] > 1:
            label = f"{label} ({counters[label]})"
        labeled_cols.append((label, col))
    return labeled_cols


def parse_building_panels(wb, sheet_name: str, fallback_label: str,
                           target_month: int, target_year: int,
                           label_aliases: dict = None) -> dict:
    """
    Generic parser for the per-building sub-meter panel sheets (Main Academic,
    Nalanda, Old Hostel, PGDM Hostel, Canteen [+Add-On], ARR, Thiruvalluvar,
    Saraswati, Substation). Each sheet lists several electrical panels with
    daily Final/Initial/Total-unit (kWh) readings; we only need Total unit.

    Some vendor sheets change their panel layout mid-sheet (a meter gets
    dropped/added and the header repeats further down with a different
    number/order of Final/Intial/Total-unit triads — e.g. the Canteen sheet
    drops "Combiowen" partway through 2026). Reading the whole sheet with only
    the FIRST header's column mapping silently attributes later rows' data to
    the wrong panel name. So we re-scan for a fresh header row every time one
    appears and re-map columns/labels from that point forward.

    label_aliases: optional {old_name: new_name} map for panels that were
    simply RENAMED at some point (same physical meter, new header text) —
    e.g. Canteen's "Combiowen" became "Sub lighting panel". Old-name readings
    are merged into the new name's series instead of appearing as a separate,
    now-permanently-zero panel.

    Extracts EVERY month present in the sheet (these logs run continuously
    since January), not just target_month — so the dashboard can offer a
    month picker for these buildings instead of only ever showing the
    current month. `days`/`panels` (current-month daily detail) are kept at
    the top level for backward compatibility with existing chart code;
    `months` holds every (year, month) found with its own days/panels/total,
    keyed by 'YYYY-MM'.

    Returns {} if the sheet/columns can't be found.
    """
    label_aliases = label_aliases or {}
    try:
        ws = find_sheet(wb, sheet_name)
    except KeyError:
        log.warning("Building sheet not found: %s", sheet_name)
        return {}

    all_rows = list(ws.iter_rows(values_only=True))

    header_rows = all_rows[:8]
    panel_cols, triad_row_idx = _find_panel_columns(header_rows)
    if not panel_cols:
        log.warning("%s: could not locate Total-unit columns", sheet_name)
        return {}

    labeled_cols = _label_columns(panel_cols, fallback_label)
    all_labels_seen = {label for label, _ in labeled_cols}

    # month_key ('YYYY-MM') -> {day -> {panel_label: value}}
    by_month = defaultdict(dict)
    row_idx = triad_row_idx  # 0-indexed position within all_rows of the current triad header
    i = triad_row_idx + 1
    while i < len(all_rows):
        row = all_rows[i]
        norm = [normalize(str(c)) for c in row if c is not None]
        # A fresh header block re-declaring further down the sheet — re-map
        # columns/labels from here on, since the layout may have changed.
        if "final" in norm and "intial" in norm:
            new_header_rows = all_rows[max(0, i - 7):i + 1]
            new_panel_cols, new_triad_offset = _find_panel_columns(new_header_rows)
            if new_panel_cols:
                labeled_cols = _label_columns(new_panel_cols, fallback_label)
                all_labels_seen |= {label for label, _ in labeled_cols}
                log.info("%s: panel layout changed at row %d -> %s",
                         sheet_name, i + 1, [l for l, _ in labeled_cols])
            i += 1
            continue

        cell_date = row[0] if row else None
        if not isinstance(cell_date, datetime):
            i += 1
            continue
        # A day with no "Final" reading entered yet (Final sits 2 cols before
        # Total-unit, same Final/Intial/Total-unit triad as Solar) isn't real
        # data — its Total-unit formula can evaluate to 0 or a large negative
        # (blank treated as 0 in the subtraction), neither of which the old
        # ">= 0 else 0" clamp could tell apart from a genuine reading. Rows
        # like this get skipped entirely instead of padded in as a 0 day, so
        # trailing not-yet-reported days don't look like real zero-consumption
        # days on the dashboard's charts.
        finals_present = [
            isinstance(row[col - 2], (int, float)) if 0 <= col - 2 < len(row) else False
            for _, col in labeled_cols
        ]
        if finals_present and not any(finals_present):
            i += 1
            continue
        month_key = f"{cell_date.year:04d}-{cell_date.month:02d}"
        day = cell_date.day
        vals = {}
        for label, col in labeled_cols:
            v = safe_num(row[col]) if col < len(row) else 0
            # Skip formula-error/meter-reset artifacts (large negative or absurd swings)
            vals[label] = v if v >= 0 else 0
        by_month[month_key][day] = vals
        i += 1

    if not by_month:
        return {}

    # Apply rename-merges: fold old-name readings into the new name's series
    # so a simple relabel doesn't look like the old meter vanished and a new
    # one appeared (see label_aliases doc above).
    if label_aliases:
        for day_values in by_month.values():
            for vals in day_values.values():
                for old_name, new_name in label_aliases.items():
                    if old_name in vals:
                        old_val = vals.pop(old_name)
                        vals[new_name] = vals.get(new_name, 0) + old_val
        all_labels_seen = {label_aliases.get(l, l) for l in all_labels_seen}

    # Union of every label ever seen (across every layout era) so a panel
    # dropped partway through the sheet (e.g. "Combiowen") still shows its
    # historical data instead of disappearing, in a stable/consistent order.
    all_labels = sorted(all_labels_seen)

    HISTORY_START_YEAR = 2024  # matches the dashboard's main Historical tab range

    months = {}
    for month_key, day_values in by_month.items():
        if int(month_key[:4]) < HISTORY_START_YEAR:
            continue
        days = sorted(day_values.keys())
        panels = [
            {"name": label, "daily": [day_values[d].get(label, 0) for d in days]}
            for label in all_labels
        ]
        total = sum(sum(p["daily"]) for p in panels)
        months[month_key] = {"days": days, "panels": panels, "total": round(total, 1)}

    target_key = f"{target_year:04d}-{target_month:02d}"
    current = months.get(target_key)
    if not current:
        # No rows at all for the target month yet — still return history.
        current = {"days": [], "panels": [{"name": label, "daily": []} for label in all_labels]}

    result = dict(current)
    result["months"] = months
    return result


# Panels that were simply renamed at some point (same physical meter, new
# header text in the vendor sheet) — keyed by BUILDING_SHEETS key, so the
# old name's readings merge into the new name instead of showing as a
# separate, permanently-zero panel. See parse_building_panels' label_aliases.
BUILDING_LABEL_ALIASES = {
    "canteen": {"Combiowen": "Sub lighting panel"},
}


def parse_all_buildings(wb, target_month: int, target_year: int) -> dict:
    """Parse every building sub-meter sheet; returns {'buildings': {key: {...}}}."""
    buildings = {}
    for key, sheet_name, label in BUILDING_SHEETS:
        aliases = BUILDING_LABEL_ALIASES.get(key)
        result = parse_building_panels(wb, sheet_name, label, target_month, target_year,
                                        label_aliases=aliases)
        if result:
            result["label"] = label
            buildings[key] = result
    if not buildings:
        return {}
    return {"buildings": buildings}


def _parse_room_meter_date(val, target_month: int, target_year: int):
    """
    Parse a date header cell from a room-meter sheet. Accepts real datetimes
    or 'DD.MM.YYYY' strings (e.g. '01.07.2026'). Returns day-of-month if it
    matches the target month/year, else None. Baseline columns like
    'up to 24/06/2026' or 'Upto 24-06-2026' are never day-columns — reject
    strings that don't parse cleanly as a single DD.MM.YYYY date.
    """
    if isinstance(val, datetime):
        if val.month == target_month and val.year == target_year:
            return val.day
        return None
    if isinstance(val, str):
        s = val.strip()
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                d = datetime.strptime(s, fmt)
                if d.month == target_month and d.year == target_year:
                    return d.day
                return None
            except ValueError:
                continue
    return None


def parse_room_meter_sheet(wb, sheet_name: str, meter_offsets: dict,
                            target_month: int, target_year: int,
                            date_header_row: int = None,
                            id_cols: tuple = (0, 1)) -> dict:
    """
    Generic parser for per-room/per-meter reading sheets (C-Block, Bramahputra)
    where each ROW is a room/meter and each COLUMN GROUP is a date, unlike the
    per-building panel sheets where rows are dates. A date's "Consumed" value
    sits at a fixed column offset from the date's header cell.

    meter_offsets: {meter_label: offset_from_date_col} e.g. {'PSSB': 1, 'LSSB': 3}
        for sheets with multiple meters per room, or {'': 1} for a single meter.
    date_header_row: 1-indexed row containing the date labels; auto-detected
        (first row with 2+ parseable target-month dates) if not given.
    id_cols: 0-indexed columns to try, in order, for the room's display name
        (falls back to the next column if the first is blank).
    Returns {'days': [...], 'panels': [{'name','daily'}]} — same shape as
    parse_building_panels so it plugs into the same BUILDINGS/UI pipeline.
    """
    try:
        ws = find_sheet(wb, sheet_name)
    except KeyError:
        log.warning("Room-meter sheet not found: %s", sheet_name)
        return {}

    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    if date_header_row is None:
        for i, row in enumerate(all_rows[:5]):
            hits = sum(1 for c in row if _parse_room_meter_date(c, target_month, target_year) is not None)
            if hits >= 1:
                date_header_row = i + 1
                break
    if date_header_row is None:
        log.warning("%s: could not locate a date header row", sheet_name)
        return {}

    header = all_rows[date_header_row - 1]
    date_cols = []  # [(day, col_idx)]
    for col_idx, cell in enumerate(header):
        day = _parse_room_meter_date(cell, target_month, target_year)
        if day is not None:
            date_cols.append((day, col_idx))
    if not date_cols:
        return {}

    panels = []
    seen_names = {}
    for row in all_rows[date_header_row:]:
        if row is None:
            continue
        name = None
        for c in id_cols:
            if c < len(row) and row[c] is not None and str(row[c]).strip():
                name = str(row[c]).strip()
                break
        if not name or normalize(name) in ("s.n", "room no", "staff name", "meter no", "month/date"):
            continue

        seen_names[name] = seen_names.get(name, 0) + 1
        label = name if seen_names[name] == 1 else f"{name} ({seen_names[name]})"

        for meter_label, offset in meter_offsets.items():
            daily = []
            for day, col_idx in date_cols:
                target_col = col_idx + offset
                v = safe_num(row[target_col]) if target_col < len(row) else 0
                daily.append(v if v >= 0 else 0)  # skip #REF!/negative meter-reset artifacts
            panel_name = f"{label} {meter_label}".strip() if meter_label else label
            panels.append({"name": panel_name, "daily": daily})

    if not panels:
        return {}
    days = sorted(d for d, _ in date_cols)
    # daily arrays above are built in date_cols order (already ascending since we scan left-to-right)
    return {"days": days, "panels": panels}


def parse_manpower(wb) -> dict:
    """
    Parse the 'Man Power' sheet: a snapshot table of staff counts by
    department and shift. Not date-indexed — reflects deployment as of the
    report date.
    Returns {} if the sheet isn't found or has no recognizable rows.
    """
    try:
        ws = find_sheet(wb, "Man Power", "Manpower")
    except KeyError:
        log.warning("Man Power sheet not found")
        return {}

    rows = list(ws.iter_rows(min_row=1, max_row=30, values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        norm = [normalize(str(c)) if c is not None else "" for c in row]
        if "first shift" in norm and "total" in norm:
            header_idx = i
            break
    if header_idx is None:
        log.warning("Man Power: header row not found")
        return {}

    header = rows[header_idx]
    col = {}
    for idx, cell in enumerate(header):
        n = normalize(str(cell)) if cell is not None else ""
        if n == "first shift":   col["first"] = idx
        elif n == "second shift": col["second"] = idx
        elif n == "third shift":  col["third"] = idx
        elif n == "general shift": col["general"] = idx
        elif n == "total":        col["total"] = idx
        elif "w & c" in n or "w&c" in n: col["wc"] = idx
        elif n == "leave":        col["leave"] = idx

    depts = []
    shift_totals = {"first": 0, "second": 0, "third": 0, "general": 0}
    wc_total = 0
    leave_total = 0
    for row in rows[header_idx + 1:]:
        name = row[0] if row else None
        if not isinstance(name, str) or not name.strip():
            continue
        label = name.strip()
        if normalize(label) == "total":
            break
        n = safe_num(row[col.get("total")]) if "total" in col else 0
        if n <= 0:
            continue
        depts.append({"name": label, "n": int(n)})
        for shift_key in ("first", "second", "third", "general"):
            if shift_key in col:
                shift_totals[shift_key] += safe_num(row[col[shift_key]])
        if "wc" in col:
            wc_total += safe_num(row[col["wc"]])
        if "leave" in col:
            leave_total += safe_num(row[col["leave"]])

    if not depts:
        log.warning("Man Power: no department rows found")
        return {}

    total_strength = sum(d["n"] for d in depts)
    active = total_strength - int(wc_total) - int(leave_total)
    return {
        "mpDepts": depts,
        "mpShifts": {
            "labels": ["First", "Second", "Third", "General"],
            "vals": [int(shift_totals["first"]), int(shift_totals["second"]),
                     int(shift_totals["third"]), int(shift_totals["general"])],
        },
        "mpTotalStrength": total_strength,
        "mpActive": active,
        "mpWc": int(wc_total),
        "mpLeave": int(leave_total),
    }


def parse_street_lights(wb) -> dict:
    """
    Parse the 'Street light Data' sheet: fixture inventory by type with
    glowing/not-glowing counts. Not date-indexed — current inventory snapshot.
    """
    try:
        ws = find_sheet(wb, "Street light Data", "Street Light Data")
    except KeyError:
        log.warning("Street light Data sheet not found")
        return {}

    rows = list(ws.iter_rows(min_row=1, max_row=60, values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        norm = [normalize(str(c)) if c is not None else "" for c in row]
        if "watts" in norm and "qty" in norm:
            header_idx = i
            break
    if header_idx is None:
        log.warning("Street light Data: header row not found")
        return {}

    header = rows[header_idx]
    col = {}
    for idx, cell in enumerate(header):
        n = normalize(str(cell)) if cell is not None else ""
        if n == "watts":     col["watts"] = idx
        elif n == "location": col["location"] = idx
        elif n == "type":     col["type"] = idx
        elif n == "qty":      col["qty"] = idx
        elif "glowing" in n and "not" not in n: col["ok"] = idx
        elif "not" in n and "glowing" in n:     col["fail"] = idx

    lights = []
    for row in rows[header_idx + 1:]:
        if not row or row[0] is None:
            continue
        try:
            int(row[0])
        except (TypeError, ValueError):
            continue
        watts = str(row[col["watts"]]).strip() if "watts" in col and row[col["watts"]] is not None else ""
        location = str(row[col["location"]]).strip() if "location" in col and row[col["location"]] is not None else ""
        qty = int(safe_num(row[col.get("qty")]))
        if qty <= 0:
            continue
        ok = int(safe_num(row[col.get("ok")]))
        fail_raw = row[col.get("fail")] if "fail" in col else None
        fail = int(safe_num(fail_raw)) if fail_raw is not None else max(0, qty - ok)
        label = f"{location.split()[0] if location else ''} {watts}".strip() or watts
        lights.append({"type": label, "total": qty, "ok": ok, "fail": fail})

    if not lights:
        log.warning("Street light Data: no fixture rows found")
        return {}

    return {"streetLights": lights}


def _parse_amc_date(val):
    """Best-effort parse of AMC date cells: datetime, 'DD.MM.YYYY' strings, or free text."""
    if isinstance(val, datetime):
        return val.strftime("%d-%b-%y")
    if isinstance(val, str):
        s = val.strip()
        if not s or s == "-":
            return "-"
        for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).strftime("%d-%b-%y")
            except ValueError:
                continue
        return s  # free text like "Oct month" — keep as-is
    return "-"


def _amc_status(next_visit_raw, target_month: int, target_year: int) -> str:
    """Derive Active/Overdue/Critical from the next-visit date vs. the report period."""
    if not isinstance(next_visit_raw, (datetime,)):
        if isinstance(next_visit_raw, str):
            for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    next_visit_raw = datetime.strptime(next_visit_raw.strip(), fmt)
                    break
                except ValueError:
                    continue
    if not isinstance(next_visit_raw, datetime):
        return "Active"  # unparseable ("-", "Dec Month") — can't judge, assume active
    report_date = datetime(target_year, target_month, 1)
    overdue_days = (report_date - next_visit_raw).days
    if overdue_days > 180:
        return "Critical"
    if overdue_days > 0:
        return "Overdue"
    return "Active"


def parse_amc(wb, target_month: int, target_year: int) -> dict:
    """
    Parse the 'AMC' sheet: vendor contract register with last/next visit dates.
    Status (Active/Overdue/Critical) is derived from how overdue next-visit is
    relative to the report month, not stored in the sheet.
    """
    try:
        ws = find_sheet(wb, "AMC")
    except KeyError:
        log.warning("AMC sheet not found")
        return {}

    rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        norm = [normalize(str(c)) if c is not None else "" for c in row]
        if "name of vendor" in norm or ("date of visit" in norm and "date of next visit" in norm):
            header_idx = i
            break
    if header_idx is None:
        log.warning("AMC: header row not found")
        return {}

    contracts = []
    for row in rows[header_idx + 1:]:
        if not row or not isinstance(row[0], str) or not row[0].strip():
            continue
        system = row[0].strip()
        last_raw = row[1] if len(row) > 1 else None
        next_raw = row[2] if len(row) > 2 else None
        vendor   = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        contracts.append({
            "system": system,
            "vendor": vendor,
            "last": _parse_amc_date(last_raw),
            "next": _parse_amc_date(next_raw),
            "status": _amc_status(next_raw, target_month, target_year),
        })

    if not contracts:
        log.warning("AMC: no contract rows found")
        return {}

    return {"amcContracts": contracts}


# ── Entry point ───────────────────────────────────────────────────────────────

def parse_electrical_file(file_path: str, target_month: int = None, target_year: int = None) -> dict:
    path = Path(file_path)
    log.info("Parsing: %s", path.name)

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    if target_month is None or target_year is None:
        try:
            ws = find_sheet(wb, "EB&DG Units", "EB&DG Units ")
        except KeyError as e:
            raise ValueError(f"Could not detect target month — 'EB&DG Units' sheet not found: {e}")
        latest = None
        for row in ws.iter_rows(min_row=5, values_only=True):
            if isinstance(row[0], datetime):
                latest = row[0]
        if latest:
            target_month = target_month or latest.month
            target_year  = target_year  or latest.year
            log.info("Auto-detected period: %s %d", latest.strftime("%B"), target_year)
        else:
            raise ValueError("Could not detect target month from Excel file — no date rows found in 'EB&DG Units'")

    result = {"month": target_month, "year": target_year}

    # Each section below is independently guarded: a renamed/missing sheet in
    # ANY one of these must not abort the whole file's parse (and therefore
    # the whole pipeline run) — it should only blank that one section while
    # everything else still comes through. See _safe_section() docstring.
    result.update(_safe_section("EB&DG Units", lambda: parse_eb_dg(wb, target_month, target_year)) or {})
    result["powerCuts"] = _safe_section("EB power cut", lambda: parse_power_cuts(wb, target_month, target_year)) or []
    result["powerCutsByMonth"] = _safe_section("EB power cut (all months)", lambda: parse_power_cuts_all_months(wb)) or {}
    result.update(_safe_section("RM Solar panel reading", lambda: parse_rm_solar(wb, target_month, target_year)) or {})
    result.update(_safe_section("Solar Panel reading", lambda: parse_main_solar(wb, target_month, target_year)) or {})

    # New electrical sub-meters
    ev_by_day = parse_ev_meter(wb, target_month, target_year)
    nescafe_by_day    = parse_tenant_meter(wb, "Nescafe coffee shop meter", target_month=target_month, target_year=target_year)
    cvb_by_day        = parse_tenant_meter(wb, "Chennai Beverage shop", target_month=target_month, target_year=target_year)
    tea_by_day        = parse_tenant_meter(wb, "Tea wheeler shop", target_month=target_month, target_year=target_year)
    yummy_by_day      = parse_tenant_meter(wb, "Yummpy's shop", target_month=target_month, target_year=target_year)
    lavasa_by_day     = parse_tenant_meter(wb, "Lavasa shop", "Lavasa Shop", target_month=target_month, target_year=target_year)
    laundry_elec      = parse_laundry_elec(wb, target_month, target_year)

    days = result.get("days", [])
    result["evDailyKWh"]       = [ev_by_day.get(d, 0)       for d in days]
    result["nescafeDailyKWh"]  = [nescafe_by_day.get(d, 0)  for d in days]
    result["cvbDailyKWh"]      = [cvb_by_day.get(d, 0)      for d in days]
    result["teaDailyKWh"]      = [tea_by_day.get(d, 0)      for d in days]
    result["yummyDailyKWh"]    = [yummy_by_day.get(d, 0)    for d in days]
    result["lavasaDailyKWh"]   = [lavasa_by_day.get(d, 0)   for d in days]
    result["laundryElecKWh"]   = [laundry_elec.get(d, 0)    for d in days]

    # Facility sections previously hardcoded in the dashboard
    result.update(parse_manpower(wb))
    result.update(parse_street_lights(wb))
    result.update(parse_amc(wb, target_month, target_year))

    # Per-building sub-meter panels (previously not surfaced anywhere)
    result.update(parse_all_buildings(wb, target_month, target_year))

    result["unmatchedSheets"] = detect_unmatched_sheets(wb)
    if result["unmatchedSheets"]:
        log.warning("Unrecognized sheet(s) in %s (not parsed): %s", path.name, result["unmatchedSheets"])
    result["sectionErrors"] = list(SECTION_ERRORS)
    SECTION_ERRORS.clear()

    wb.close()
    log.info("Parsing complete. Days: %s", result.get("days", []))
    return result


CBLOCK_SHEETS = [
    ("cBlock",      "C-block Reading",     "C-Block",      {"PSSB": 1, "LSSB": 3}, (1, 2)),
    ("bramahputra", "Bramahputra Reading", "Bramahputra",  {"": 1},                (1, 2)),
]


def parse_cblock_file(file_path: str, target_month: int = None, target_year: int = None) -> dict:
    """
    Parse the 'C Block Panel Reading' Excel file — a separate email attachment
    (not part of the main Electrical/STP report) covering C-Block guest-house
    rooms and the Bramahputra block meters. Feeds into the same BUILDINGS
    structure the dashboard already renders (see parse_all_buildings), so no
    new UI wiring is needed beyond adding these two keys.
    """
    path = Path(file_path)
    log.info("Parsing C-Block file: %s", path.name)

    wb = openpyxl.load_workbook(str(path), read_only=False, data_only=True)

    if target_month is None or target_year is None:
        try:
            ws = find_sheet(wb, "C-block Reading", "C-Block Reading")
        except KeyError as e:
            raise ValueError(f"C-Block file has no recognizable main sheet: {e}")
        latest = None
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for cell in row:
                if isinstance(cell, datetime):
                    latest = cell if latest is None or cell > latest else latest
                elif isinstance(cell, str):
                    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
                        try:
                            dt = datetime.strptime(cell.strip(), fmt)
                            latest = dt if latest is None or dt > latest else latest
                            break
                        except ValueError:
                            continue
        if latest:
            target_month = target_month or latest.month
            target_year  = target_year  or latest.year
            log.info("C-Block auto-detected period: %s %d", latest.strftime("%B"), target_year)
        else:
            raise ValueError("Could not detect month from C-Block Excel file")

    buildings = {}
    for key, sheet_name, label, meter_offsets, id_cols in CBLOCK_SHEETS:
        result = parse_room_meter_sheet(wb, sheet_name, meter_offsets, target_month, target_year,
                                         id_cols=id_cols)
        if result:
            result["label"] = label
            buildings[key] = result

    unmatched_sheets = detect_unmatched_sheets(wb)
    if unmatched_sheets:
        log.warning("Unrecognized sheet(s) in %s (not parsed): %s", path.name, unmatched_sheets)

    wb.close()
    if not buildings:
        log.warning("C-Block file: no data parsed from either sheet")
        return {"unmatchedSheets": unmatched_sheets}
    return {"month": target_month, "year": target_year, "buildings": buildings, "unmatchedSheets": unmatched_sheets}


def _find_hall_columns(header_rows: list) -> list:
    """
    Locate each hall's sub-columns in a Chair Count sheet. Row layout:
      row 0: hall names spanning several columns each (merged cells)
      row 1: (blank, or occupant/notes row — ignored)
      row 2: sub-header — W.CH / F.Ch / Pad.Ch / Total per hall (col count varies)
    Returns [(hall_name, {'wch':col,'fch':col,'padch':col,'total':col})], using
    whichever sub-columns actually exist for that hall (F.Ch is optional).
    """
    sub_header_idx = None
    for i, row in enumerate(header_rows):
        norm = [normalize(str(c)) for c in row if c is not None]
        if any("w.ch" in n or n == "wch" for n in norm):
            sub_header_idx = i
            break
    if sub_header_idx is None:
        return [], None

    # The hall-name row sits directly above the W.CH/Pad.Ch/Total sub-header —
    # NOT necessarily row 0, which may just be a sheet title banner.
    name_row = header_rows[sub_header_idx - 1] if sub_header_idx >= 1 else header_rows[0]
    sub_row = header_rows[sub_header_idx]

    # Map each column to its hall name (forward-fill across the merged span).
    hall_by_col = {}
    last_name = None
    for col_idx, cell in enumerate(name_row):
        if cell is not None and str(cell).strip():
            last_name = str(cell).strip()
        if col_idx >= 2:  # skip S.No./Date columns
            hall_by_col[col_idx] = last_name

    halls = {}  # name -> {sub_label: col}
    for col_idx, cell in enumerate(sub_row):
        if cell is None:
            continue
        label = normalize(str(cell))
        hall_name = hall_by_col.get(col_idx)
        if not hall_name:
            continue
        halls.setdefault(hall_name, {})
        if "total" in label:
            halls[hall_name]["total"] = col_idx
        elif "w.ch" in label or label == "wch":
            halls[hall_name]["wch"] = col_idx
        elif "f.ch" in label or label == "fch":
            halls[hall_name]["fch"] = col_idx
        elif "pad" in label:
            halls[hall_name]["padch"] = col_idx

    result = [(name, cols) for name, cols in halls.items() if "total" in cols]
    return result, sub_header_idx


CHAIR_COUNT_SHEETS = [
    ("hall1to6", "Hall 1-6", "Hall 1-6"),
    ("nalandaArr", "Nalanda & ARR", "Nalanda & ARR"),
]


def parse_chair_count_sheet(wb, sheet_name: str) -> dict:
    """
    Parse one sheet of the Chair Count Details file. Each row is a date,
    each hall has W.CH (wheelchair) / F.Ch (folding) / Pad.Ch (padded) /
    Total chair columns. This is inventory, not daily consumption — counts
    rarely change — so we report the LATEST reading per hall plus a short
    recent-history trend for context, rather than summing.
    Returns {'halls': [{'name','wch','fch','padch','total'}], 'asOf': 'DD-MM-YYYY'}
    """
    try:
        ws = find_sheet(wb, sheet_name)
    except KeyError:
        log.warning("Chair count sheet not found: %s", sheet_name)
        return {}

    header_rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    hall_cols, sub_header_idx = _find_hall_columns(header_rows)
    if not hall_cols:
        log.warning("%s: could not locate hall/Total columns", sheet_name)
        return {}

    data_start = sub_header_idx + 2  # 1-indexed row right after the sub-header row
    latest_date = None
    latest_row = None
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        cell_date = row[1] if len(row) > 1 else None
        if not isinstance(cell_date, datetime):
            continue
        # Skip rows where every hall's Total is blank/zero (not yet filled in)
        if not any(safe_num(row[cols["total"]]) > 0 for _, cols in hall_cols if cols["total"] < len(row)):
            continue
        if latest_date is None or cell_date >= latest_date:
            latest_date = cell_date
            latest_row = row

    if latest_row is None:
        return {}

    halls = []
    for name, cols in hall_cols:
        total_col = cols.get("total")
        if total_col is None or total_col >= len(latest_row):
            continue
        halls.append({
            "name": name,
            "wch": int(safe_num(latest_row[cols["wch"]])) if "wch" in cols and cols["wch"] < len(latest_row) else 0,
            "fch": int(safe_num(latest_row[cols["fch"]])) if "fch" in cols and cols["fch"] < len(latest_row) else 0,
            "padch": int(safe_num(latest_row[cols["padch"]])) if "padch" in cols and cols["padch"] < len(latest_row) else 0,
            "total": int(safe_num(latest_row[total_col])),
        })

    return {"halls": halls, "asOf": latest_date.strftime("%d-%b-%Y")}


def parse_chair_count_file(file_path: str) -> dict:
    """
    Parse the 'Chair Count Details' Excel file — a separate email attachment
    tracking classroom/hall chair inventory (not energy/water). Returns
    {'chairCounts': {key: {'label','halls','asOf'}}}.
    """
    path = Path(file_path)
    log.info("Parsing Chair Count file: %s", path.name)

    wb = openpyxl.load_workbook(str(path), data_only=True)

    chair_counts = {}
    for key, sheet_name, label in CHAIR_COUNT_SHEETS:
        result = parse_chair_count_sheet(wb, sheet_name)
        if result:
            result["label"] = label
            chair_counts[key] = result

    wb.close()
    if not chair_counts:
        log.warning("Chair Count file: no data parsed from either sheet")
        return {}
    return {"chairCounts": chair_counts}


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")

    elec_file = BASE_DIR / "downloads" / "Electrical, AC, Carpentry, STP & Plumbing Daily Report June 2026.xlsx"
    data = parse_electrical_file(str(elec_file))

    print("\n=== PARSED DATA SUMMARY ===")
    print(f"Period     : {data.get('month')}/{data.get('year')}")
    print(f"Days found : {data.get('days', [])}")
    print(f"junEB      : {data.get('junEB', [])[:5]} ...")
    print(f"junDG1     : {data.get('junDG1', [])[:5]} ...")
    print(f"junDiesel  : {data.get('junDiesel', [])[:5]} ...")
    print(f"junStock   : {data.get('junStock', [])[:5]} ...")
    print(f"Power cuts : {len(data.get('powerCuts', []))} events")
    print(f"RM Solar   : {data.get('rmSolarDaily', {}).get('canteen', [])[:5]} ...")
